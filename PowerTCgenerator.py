import os
import re
import io
import shutil
from openpyxl import load_workbook
import openpyxl
import xml.etree.ElementTree as ET
from xml.dom import minidom

"""
*****************************************
     Author: Alon Gritsovsky
     Welcome to Power TC generator
*****************************************
"""

print("***********************************************")
print("****** Welcome to Power TC generator  *********")
print("***********************************************")

excel_files = {
    "1": "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\All_power_tests_for_1250.xlsx",
    "2": "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\All_power_tests_for_1350.xlsx"
}

print("Choose an Excel file:")
for key, file_path in excel_files.items():
    print(f"{key}. {os.path.basename(file_path)}")

choice = input("Enter your choice (1 or 2): ")

file_path = excel_files.get(choice)
if not file_path:
    print("Invalid choice. Exiting...")
    exit(1)

# Load the chosen Excel file
print("Loading Excel file please wait...")
workbook = load_workbook(filename=file_path, data_only=True)
sheet_names = workbook.sheetnames

# Prompt the user to choose a sheet
print("Choose a sheet:")
for i, sheet_name in enumerate(sheet_names):
    print(f"{i + 1}. {sheet_name}")

sheet_choice = input("Enter your choice (1 to {0}): ".format(len(sheet_names)))

# Validate the user's sheet choice
try:
    sheet_index = int(sheet_choice) - 1
    if 0 <= sheet_index < len(sheet_names):
        selected_sheet = workbook[sheet_names[sheet_index]]
    else:
        raise ValueError
except ValueError:
    print("Invalid choice. Exiting...")
    exit(1)

output_campaign = f"C:\\QA\\autotester_tools\\AutoTester_Campaigns\\{sheet_names[sheet_index]}"
output_tests = f"C:\\QA\\autotester_tools\\AutoTester_TestCases\\{sheet_names[sheet_index]}\\{sheet_names[sheet_index]}"
output_procedures = f"C:\\QA\\autotester_tools\\AutoTester_Procedures\\{sheet_names[sheet_index]}\\{sheet_names[sheet_index]}"
campaign_name = f"{sheet_names[sheet_index]}.atc"
sheet_name_selected = sheet_names[sheet_index]

"""
this function checks if the given campaign (choiced sheet from the user) exists or not
if yes its ask to overwrite it, if yes return true if not return false.
"""


def check_campaign_exists(output_campaign, output_tests, output_procedures, campaign_name):
    if os.path.exists(output_campaign) and os.path.exists(output_tests) and os.path.exists(output_procedures):
        campaign_name_checker = os.path.join(output_campaign, campaign_name)
        if os.path.isfile(campaign_name_checker):
            response = input("This campaign already exists. Do you want to overwrite it? (y/n): ")
            if response.lower() == "y":
                return True
            elif response.lower() == "n":
                return False
            else:
                print("Invalid response. Please enter 'y' or 'n'.")
                return check_campaign_exists(output_campaign, output_tests, output_procedures, campaign_name)


default_tests = {
    'Power_Measurement_CATM_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Measurement_Tests_CATM_1350.attc",
    'Power_Functional_CATM_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Functional_Tests_CATM_1350.attc",
    'Power_Measurement_CATM_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Measurement_Tests_CATM_1250.attc",
    'Power_Functional_CATM_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Functional_Tests_CATM_1250.attc",
    'Power_Measurement_NB-IOT_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Measurement_Tests_NB-IOT_1350.attc",
    'Power_Functional_NB-IOT_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Functional_Tests_NB-IOT_1350.attc",
    'Power_Measurement_NB-IOT_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Measurement_Tests_NB-IOT_1250.attc",
    'Power_Functional_NB-IOT_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Tests\\Functional_Tests_NB-IOT_1250.attc"
}

default_procedures = {
    'Power_Measurement_CATM_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1350",
    'Power_Functional_CATM_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1350",
    'Power_Measurement_CATM_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1250",
    'Power_Functional_CATM_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1250",
    'Power_Measurement_NB-IOT_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1350",
    'Power_Functional_NB-IOT_1350': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1350",
    'Power_Measurement_NB-IOT_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1250",
    'Power_Functional_NB-IOT_1250': "C:\\QA\\autotester_tools\\AutoTester_Scripts\\PowerTCgenerator\\Defaults\\Default_Procedures\\procedures_for_1250"
}


def replace_measurement_variables(variables_list, row):
    for item in variables_list:
        name = item['name']
        if name == 'edrx_cycle':
            item['initialValue'] = 'none' if str(row[6]) == 'NA' else str(row[6])
            item['initialValue_child'] = item['initialValue']
        elif name == 'Tx_power':
            if str(row[12]) == 'NA':
                item['initialValue'] = 'none'
                item['initialValue_child'] = 'none'
            else:
                item['initialValue'] = str(row[12])
                item['initialValue_child'] = str(row[12])
        elif name == 'expected_AVG':
            item['initialValue'] = str(row[25])
            item['initialValue_child'] = str(row[25])
        elif name == 'expected_floor':
            item['initialValue'] = str(row[24])
            item['initialValue_child'] = str(row[24])
        elif name == 'ptw':
            if str(row[7]) == 'NA':
                item['initialValue'] = 'none'
                item['initialValue_child'] = 'none'
            else:
                item['initialValue'] = str(row[7])
                item['initialValue_child'] = str(row[7])
        elif name == 'tput_case':
            item['initialValue'] = str(row[18])
            item['initialValue_child'] = str(row[18])
        elif name == 'Logs':
            item['initialValue'] = str(row[20])
            item['initialValue_child'] = str(row[20])
        elif name == 'NW_mode':
            if str(row[4]) == 'eDRX_eDRX' or str(row[4]) == 'iDRX_eDRX':
                item['initialValue'] = 'eDRX'
                item['initialValue_child'] = 'eDRX'
            elif str(row[4]) == 'iDRX_iDRX':
                item['initialValue'] = 'iDRX'
                item['initialValue_child'] = 'iDRX'
            else:
                item['initialValue'] = str(row[4])
                item['initialValue_child'] = str(row[4])
        elif name == 'target_Temperature':
            item['initialValue'] = f"{str(row[10])}.0"
            item['initialValue_child'] = item['initialValue']
        elif name == 'PagingCycle':
            if str(row[8]) == 'NA':
                item['initialValue'] = 'none'
                item['initialValue_child'] = 'none'
            else:
                item['initialValue'] = str(row[8])
                item['initialValue_child'] = str(row[8])
        elif name == 'BW':
            item['initialValue'] = str(row[11])
            item['initialValue_child'] = str(row[11])
        elif name == 'scenario':
            item['initialValue'] = str(row[3])
            item['initialValue_child'] = str(row[3])
        elif name == 'Voltage':
            item['initialValue'] = str(row[9])
            item['initialValue_child'] = str(row[9])
        elif name == 'target_SINR':
            item['initialValue'] = str(row[17])
            item['initialValue_child'] = str(row[17])
        elif name == 'band':
            item['initialValue'] = str(row[14])
            item['initialValue_child'] = str(row[14])
        elif name == 'RTOS_state':
            item['initialValue'] = str(row[13])
            item['initialValue_child'] = str(row[13])
        elif name == 'cycle_to_measure':
            item['initialValue'] = str(row[15])
            item['initialValue_child'] = str(row[15])
        elif name == 'Tput':
            if sheet_name_selected in ['Power_Measurement_CATM_1350', 'Power_Measurement_CATM_1250']:
                if str(row[18]) != 'NA' and str(row[19]) != 'NA':
                    Tput = f"ul{str(row[18])}/dl{str(row[19])}"
                    item['initialValue'] = Tput
                    item['initialValue_child'] = Tput
        elif name == 'target_RSRP':
            item['initialValue'] = str(row[16])
            item['initialValue_child'] = str(row[16])
        elif name == 'sleep_mode':
            item['initialValue'] = str(row[5])
            item['initialValue_child'] = str(row[5])
        elif name == 'use_2_cell_configuration':
            item['initialValue'] = str(row[21])
            item['initialValue_child'] = str(row[21])
        elif name == 'cell_search_mechanism':
            item['initialValue'] = str(row[22])
            item['initialValue_child'] = str(row[22])


def replace_functional_variables(variables_list, row):
    for item in variables_list:
        name = item['name']
        if name == 'edrx_cycle':
            item['initialValue'] = 'none' if str(row[6]) == 'NA' else str(row[6])
            item['initialValue_child'] = item['initialValue']
        elif name == 'Logs':
            item['initialValue'] = str(row[16])
            item['initialValue_child'] = str(row[16])
        elif name == 'NW_mode':
            if str(row[4]) == 'eDRX_eDRX' or str(row[4]) == 'iDRX_eDRX':
                item['initialValue'] = 'eDRX'
                item['initialValue_child'] = 'eDRX'
            elif str(row[4]) == 'iDRX_iDRX':
                item['initialValue'] = 'iDRX'
                item['initialValue_child'] = 'iDRX'
            else:
                item['initialValue'] = str(row[4])
                item['initialValue_child'] = str(row[4])
        elif name == 'target_Temperature':
            item['initialValue'] = f"{str(row[10])}.0"
            item['initialValue_child'] = item['initialValue']
        elif name == 'ptw':
            item['initialValue'] = 'none' if str(row[7]) == 'NA' else str(row[7])
            item['initialValue_child'] = item['initialValue']
        elif name == 'PagingCycle':
            if str(row[8]) == 'NA':
                item['initialValue'] = 'none'
                item['initialValue_child'] = 'none'
            else:
                item['initialValue'] = str(row[8])
                item['initialValue_child'] = str(row[8])
        elif name == 'BW':
            item['initialValue'] = str(row[11])
            item['initialValue_child'] = str(row[11])
        elif name == 'scenario':
            item['initialValue'] = str(row[3])
            item['initialValue_child'] = str(row[3])
        elif name == 'Voltage':
            item['initialValue'] = str(row[9])
            item['initialValue_child'] = str(row[9])
        elif name == 'target_SINR':
            item['initialValue'] = str(row[15])
            item['initialValue_child'] = str(row[15])
        elif name == 'iteration':
            item['initialValue'] = str(row[17])
            item['initialValue_child'] = str(row[17])
        elif name == 'interval':
            item['initialValue'] = str(row[18])
            item['initialValue_child'] = str(row[18])
        elif name == 'band':
            item['initialValue'] = str(row[13])
            item['initialValue_child'] = str(row[13])
        elif name == 'RTOS_state':
            item['initialValue'] = str(row[12])
            item['initialValue_child'] = str(row[12])
        elif name == 'target_RSRP':
            item['initialValue'] = str(row[14])
            item['initialValue_child'] = str(row[14])
        elif name == 'sleep_mode':
            item['initialValue'] = str(row[5])
            item['initialValue_child'] = str(row[5])


"""
this function creates the campaign and the tests
"""


def create_campaign_and_tests(output_campaign, output_tests, output_procedures, campaign_name):
    default_test = default_tests.get(sheet_name_selected)
    default_procedures_folder = default_procedures.get(sheet_name_selected)
    for filename in os.listdir(default_procedures_folder):
        if filename.endswith('.atap'):
            source_file = os.path.join(default_procedures_folder, filename)
            destination_file = os.path.join(output_procedures, filename)
            shutil.copy(source_file, destination_file)
    for filename in os.listdir(output_procedures):
        if filename.endswith('.atap'):
            file_path = os.path.join(output_procedures, filename)
            tree = ET.parse(file_path)
            root = tree.getroot()
            root.set('path', file_path)
            elements = root.findall(".//*[@path]")
            for element in elements:
                current_path = element.get('path')
                current_filename = os.path.basename(current_path)
                new_filepath = os.path.join(output_procedures, current_filename)
                element.set('path', new_filepath)
            tree.write(file_path, encoding='utf-8', xml_declaration=True, method='xml')
    if not default_test:
        print("Invalid sheet name. Exiting...")
        exit(1)
    if not default_procedures_folder:
        print("Invalid sheet name. Exiting...")
        exit(1)

    for row in selected_sheet.iter_rows(min_row=2, values_only=True):
        File_name = row[1]
        output_testCase = f'{output_tests}\\{File_name}.attc'
        shutil.copy(default_test, output_testCase)
        tree = ET.parse(output_testCase)
        root = tree.getroot()
        root.set('name', row[0])
        root.set('description', row[1])
        root.set('specCategory', row[2])
        elements = root.findall(".//*[@path]")
        for element in elements:
            path_value = element.get("path")
            file_path, file_name = os.path.split(path_value)
            new_file_path = os.path.join(output_procedures, file_name)
            element.set("path", new_file_path)
        variables_element = root.find('variables')
        variables_list = []
        for variable_element in variables_element.iter('variable'):
            variable_dict = {
                'name': variable_element.get('name'),
                'initialValue': variable_element.get('initialValue'),
                'initialValue_child': variable_element.find('initialValue').text}
            variables_list.append(variable_dict)

        # Replace variables in the measurement tests
        if sheet_name_selected in ['Power_Measurement_CATM_1350', 'Power_Measurement_CATM_1250',
                                   'Power_Measurement_NB-IOT_1350', 'Power_Measurement_NB-IOT_1250']:
            replace_measurement_variables(variables_list, row)
        # Replace variables in the functional tests
        elif sheet_name_selected in ['Power_Functional_CATM_1350', 'Power_Functional_CATM_1250',
                                     'Power_Functional_NB-IOT_1350', 'Power_Functional_NB-IOT_1250']:
            replace_functional_variables(variables_list, row)

        for variable_element, variable_dict in zip(variables_element.iter('variable'), variables_list):
            variable_element.set('initialValue', variable_dict['initialValue'])
            variable_element.find('initialValue').text = variable_dict['initialValue_child']
        for variable in root.iter('variable'):
            if variable.get('type') == 'Number':
                initial_value = variable.find('initialValue')
                if initial_value is not None:
                    value = initial_value.text
                    if value == 'None':
                        initial_value.text = '0.0'
                    elif value is not None and '.' not in value:
                        initial_value.text = value + '.0'
                    attr_value = variable.get('initialValue')
                    if attr_value == 'None':
                        variable.set('initialValue', '0.0')
                    elif attr_value is not None and '.' not in attr_value:
                        variable.set('initialValue', attr_value + '.0')
        tree.write(output_testCase, encoding='utf-8', xml_declaration=True)

    """  Create the Campaign  """

    campaign = ET.Element("campaign", name=sheet_names[sheet_index], description="", randomRunOrder="false")
    ParserRevision = ET.SubElement(campaign, "ParserRevision").text = "1.0"
    testCases = ET.SubElement(campaign, "testCases")
    test_cases_values = [row[0] for row in selected_sheet.iter_rows(min_row=2, values_only=True)]
    test_descriptions = [row[1] for row in selected_sheet.iter_rows(min_row=2, values_only=True)]
    file_paths = [os.path.join(output_tests, file_name) for file_name in os.listdir(output_tests) if
                  os.path.isfile(os.path.join(output_tests, file_name))]

    for test_case, description in zip(test_cases_values, test_descriptions):
        matching_paths = [path for path in file_paths if os.path.basename(path).startswith(description)]
        if matching_paths:
            testCase = ET.SubElement(testCases, "testCase")
            testCase.set("name", "%s" % test_case)
            testCase.set("description", "%s" % description)
            testCase.set("file", "%s" % matching_paths[0])
            testCase.set("author", "Alon Gritsovsky")
            executionControl = ET.SubElement(testCase, "executionControl")
            run2 = ET.SubElement(executionControl, "run")
            run2.text = "true"
            timeout1 = ET.SubElement(executionControl, "timeout")
            timeout1.set("timeUnit", "SECONDS")
            timeout1.text = "0"
            iterationsLimit1 = ET.SubElement(executionControl, "iterationsLimit")
            iterationsLimit1.text = "1.0"
            maxRuns1 = ET.SubElement(executionControl, "maxRuns")
            maxRuns1.text = "0"
            repeatCondition1 = ET.SubElement(executionControl, "repeatCondition")
            repeatCondition1.text = "NONE"

    tree = ET.ElementTree(campaign)
    with open(f"{output_campaign}\\{campaign_name}", "w", encoding="utf-8") as xml_file:
        tree_str = ET.tostring(tree.getroot(), encoding="utf-8")
        parsed_tree = minidom.parseString(tree_str)
        pretty_xml_str = parsed_tree.toprettyxml(indent="    ")
        xml_file.write(pretty_xml_str)


while True:
    if not os.path.exists(output_campaign) and not os.path.exists(output_tests) and not os.path.exists(
            output_procedures):
        os.mkdir(output_campaign)
        first_path_tests = f"C:\\QA\\autotester_tools\\AutoTester_TestCases\\{sheet_names[sheet_index]}"
        os.mkdir(first_path_tests)
        output_tests = f"C:\\QA\\autotester_tools\\AutoTester_TestCases\\{sheet_names[sheet_index]}\\{sheet_names[sheet_index]}"
        os.mkdir(output_tests)
        first_path_procedures = f"C:\\QA\\autotester_tools\\AutoTester_Procedures\\{sheet_names[sheet_index]}"
        os.mkdir(first_path_procedures)
        output_procedures = f"C:\\QA\\autotester_tools\\AutoTester_Procedures\\{sheet_names[sheet_index]}\\{sheet_names[sheet_index]}"
        os.mkdir(output_procedures)
        print("Generating campaign please wait...")
        create_campaign_and_tests(output_campaign, output_tests, output_procedures, campaign_name)
        break
    if check_campaign_exists(output_campaign, output_tests, output_procedures, campaign_name):
        campaign_to_remove = os.path.join(output_campaign, campaign_name)
        os.remove(campaign_to_remove)
        shutil.rmtree(output_tests)
        os.mkdir(output_tests)
        shutil.rmtree(output_procedures)
        os.mkdir(output_procedures)
        print("Generating campaign please wait...")
        create_campaign_and_tests(output_campaign, output_tests, output_procedures, campaign_name)
        break
    else:
        while True:
            new_campaign_name = input("Enter a new name for the campaign: ")
            new_output_tests = f"C:\\QA\\autotester_tools\\AutoTester_TestCases\\{sheet_names[sheet_index]}\\{new_campaign_name}"
            new_output_procedures = f"C:\\QA\\autotester_tools\\AutoTester_Procedures\\{sheet_names[sheet_index]}\\{new_campaign_name}"

            if os.path.exists(new_output_tests) or os.path.exists(new_output_procedures):
                print("Campaign with that name already exists. Please enter a different name.")
            else:
                new_campaign_name_file = new_campaign_name + ".atc"
                os.mkdir(new_output_tests)
                os.mkdir(new_output_procedures)
                print("Generating campaign please wait...")
                create_campaign_and_tests(output_campaign, new_output_tests, new_output_procedures,
                                          new_campaign_name_file)
                break
        break
